#!/usr/bin/env python3
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
import sys

HEADERS = {'Authorization' : 'Bearer key1agtUnabRLb2LS', 'accept' : 'text/plain'}
BASE_URL = 'https://api.airtable.com/v0/appj3UWymNh6FgtGR/'
VIEW = 'view=Grid%20view'

# Values related to the place table
PLACE_TABLE_NAME = 'Help%20Services'
PLACE_TABLE_VAR = 'placesTableCached'
PLACE_TABLE_MAP = {
    'name' : 'Name',
    'name_es' : 'Name-ES',
    'catSubcatId' : 'CatSubcat',
    'city' : 'City',
    'category' : 'Category',
    'subcategory' : 'Subcategory',
    'phone' : 'Phone Number',
    'address' : 'Physical Address',
    'latitude' : 'Latitude',
    'longitude' : 'Longitude',
    'url' : 'Web address',
    'email' : 'Email Address',
    'hours' : 'Hours of operation',
    'hours_es' : 'Hours of operation-ES',
    'description' : 'Description',
    'description_es' : 'Description-ES',
    'wheelchair' : 'Wheelchair access (y)',
    'languageHelp' : 'Language Help (y)',
    'schedule' : 'schedule',
}

# Values related to the Category table
CATEGORY_TABLE_NAME = 'Categories'
CATEGORY_TABLE_VAR = 'categoryTableCached'
CATEGORY_TABLE_MAP = {
      'name' : 'Name',
      'name_es' : 'Name-ES',
      'subcategories' : 'Subcategories',
}

# Values related to the Subcategory table
SUBCATEGORY_TABLE_NAME = 'Subcategories'
SUBCATEGORY_TABLE_VAR = 'subcategoryTableCached'
SUBCATEGORY_TABLE_MAP = {
      'categoryId' : 'Category',
      'name' : 'Name',
      'name_es' : 'Name-ES',
}

# Values related to the CatSubcat table
CATSUBCAT_TABLE_NAME = 'CatSubcats'
CATSUBCAT_TABLE_VAR = 'catSubcatTableCached'
# Cat Subcat handled specially, so doesn't use table map.
CATSUBCAT_TABLE_MAP = {}

# Values related to the Cities table
CITY_TABLE_NAME = 'Cities'
CITY_TABLE_VAR = 'cityTableCached'
CITY_TABLE_MAP = {
      'name' : 'Name',
}

ALERT_TABLE_NAME = 'Alerts'
ALERT_TABLE_VAR = 'alertTableCached'
ALERT_TABLE_MAP = {
    'title' : 'Title',
    'displayDate' : 'Display Date',
    'startDate' : 'StartDate',
    'endDate' : 'EndDate',
    'note' : 'Notes',
}

# Make a record in our desired output format
def make_record(record_in, key_pairs):
    record_out = {}
    record_out["id"] = record_in["id"]
    for key in key_pairs:
        new_key = key
        old_key = key_pairs[new_key]
        record_out[new_key] = ''
        if old_key in record_in['fields']:
            record_out[new_key] = record_in['fields'][old_key]
    return record_out

# Make a record in our desired output format -- special handling for 
# special CatSubcat features
def make_record_catsubcat(record_in):
    catSubcatId = record_in["id"]
    catSubcatName = record_in['fields']['Name']
    categoryId = record_in['fields']['Category'][0]

    subcategoryId = ''
    subcategoryName = ''
    subcategoryNameSpanish = ''
    if 'Subcategory' in record_in['fields']:
        subcategoryId = record_in['fields']['Subcategory'][0]
        subcategoryName = record_in['fields']['SubcategoryString'][0]
        subcategoryNameSpanish = record_in['fields']['Subcategory-ES'][0]

    record_out = {
      'catSubcatId': catSubcatId,
      'catSubcatName': catSubcatName,
      'categoryId' : categoryId,
      'subcategoryId' : subcategoryId,
      'name' : subcategoryName,
      'name_es' : subcategoryNameSpanish,
      'places' : []
    }
    return record_out

# Use map to convert a table to the form we want. Note the special code for the 
# CatSubcat table because it is structured a little different from the others.
def table_map(table_name, table_raw, key_pairs):
    if table_name == CATSUBCAT_TABLE_NAME:
        table = map(lambda record: make_record_catsubcat(record), table_raw)
    else:
        table = map(lambda record: make_record(record, key_pairs), table_raw)
    table = list(table)
    return table

# Air table only allows pages of up to 100 records at a time. This gets a page
def get_page(url):
    r = requests.get(url, headers=HEADERS)
    data = r.json()
    records = data['records']
    offset = None
    if 'offset' in data:
        offset = data['offset']
    return records, offset

def get_table(table_name, mapping):
    # Get the table data from air table
    table_url = BASE_URL+table_name
    page, offset = get_page(table_url+'?'+VIEW)
    table_raw = page
    # Air table only allows 100 records at a time, so loop to get them all
    while offset:
        page, offset = get_page(table_url+'?offset='+offset+'&'+VIEW)
        table_raw.extend(page)
    # Map it into the form we need
    table = table_map(table_name, table_raw, mapping)
    return table

def do_table(table_name, mapping, var_name, f):
    table = get_table(table_name, mapping)
    # Write it into the javascript file
    print('const', var_name, '=', table, ';', file=f)

def do_mailmerge(dir, place_table, category_table, catsubcat_table, language_str):
    mailmerge_table = []
    for record in place_table:
        catsubcats = record['catSubcatId']
        for cat_subcat_id in catsubcats:
            # Wrangle the cats and subcats from the catSubcat table
            catsubcat_record = list(filter(lambda catsubcat_rec: catsubcat_rec['catSubcatId'] == cat_subcat_id, catsubcat_table))
            # Assuming the above returns exactly one record
            category_id = catsubcat_record[0]['categoryId']
            category_record = list(filter(lambda cat_rec: cat_rec['id'] == category_id, category_table))
            # Assuming the above returns exactly one record
            category_str = category_record[0]['name'+language_str]
            subcategory_str = catsubcat_record[0]['name'+language_str]

            mailmerge_table.append({
                'Category' : category_str,
                'Subcategory' : subcategory_str,
                'Service Name' : record['name'],
                'Phone Number' : record['phone'],
                'Physical Address' : record['address'],
                'Hours of operation' : record['hours'],
                'Description' : record['description'+language_str],
                'Wheelchair access (y)' : record['wheelchair'],
                'Language Help (y)' : record['languageHelp'],
                'Web address' : record['url'],
                'Email Address' : record['email'],
            })

    mailmerge_table = sorted(mailmerge_table, key = lambda i: (i['Category'], i['Subcategory'], i['Service Name']))

    # Save as csv file
    csv_columns = list(mailmerge_table[0].keys())
    filename = dir+'final_book'+language_str
    csv_file = filename+'.csv'
    xls_file = filename+'.xls'
    try:
        with open(csv_file, 'w') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=csv_columns)
            writer.writeheader()
            for data in mailmerge_table:
                writer.writerow(data)
    except IOError:
        print("I/O error")
        exit

    # Convert the csv file to an Excel file
    f = open(csv_file, 'r')
    reader = csv.reader(f)
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "Sheet1"

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws[column_letter+str(row_index+1)] = cell

    wb.save(filename = xls_file)

dir = './'
if len(sys.argv) > 1:
    dir = sys.argv[1]
    if dir[-1] != '/':
        dir += '/'
# Process each table
f = open(dir+'cachedInlineTables.js', 'w')
do_table(PLACE_TABLE_NAME, PLACE_TABLE_MAP, PLACE_TABLE_VAR, f)
do_table(CATEGORY_TABLE_NAME, CATEGORY_TABLE_MAP, CATEGORY_TABLE_VAR, f)
do_table(SUBCATEGORY_TABLE_NAME, SUBCATEGORY_TABLE_MAP, SUBCATEGORY_TABLE_VAR, f)
do_table(CATSUBCAT_TABLE_NAME, CATSUBCAT_TABLE_MAP, CATSUBCAT_TABLE_VAR, f)
do_table(CITY_TABLE_NAME, CITY_TABLE_MAP, CITY_TABLE_VAR, f)
do_table(ALERT_TABLE_NAME, ALERT_TABLE_MAP, ALERT_TABLE_VAR, f)

place_table = get_table(PLACE_TABLE_NAME, PLACE_TABLE_MAP)
category_table = get_table(CATEGORY_TABLE_NAME, CATEGORY_TABLE_MAP)
catsubcat_table = get_table(CATSUBCAT_TABLE_NAME, CATSUBCAT_TABLE_MAP)
do_mailmerge(dir, place_table, category_table, catsubcat_table, '')
do_mailmerge(dir, place_table, category_table, catsubcat_table, '_es')